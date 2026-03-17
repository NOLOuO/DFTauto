import argparse
import csv
import json
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from init_pipeline import build_paths


def load_json_files(input_dirs: list[Path]) -> list[dict]:
    papers = []
    for input_dir in input_dirs:
        papers.extend(
            json.loads(path.read_text(encoding="utf-8"))
            for path in sorted(input_dir.glob("*.json"))
        )
    return papers


def load_csv_rows(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def parse_note_sections(note_path: Path) -> dict:
    if not note_path.exists():
        return {}
    content = note_path.read_text(encoding="utf-8")
    sections = {}
    current = None
    buffer = []
    for line in content.splitlines():
        if line.startswith("## "):
            if current is not None:
                sections[current] = "\n".join(buffer).strip()
            current = line[3:].strip()
            buffer = []
        else:
            buffer.append(line)
    if current is not None:
        sections[current] = "\n".join(buffer).strip()
    return sections


def build_workbook(papers: list[dict], queue_rows: list[dict], concept_map: dict, output_path: Path) -> None:
    wb = Workbook()
    ws_papers = wb.active
    ws_papers.title = "papers"
    ws_papers.append(
        [
            "title",
            "doi",
            "authors",
            "section_count",
            "reference_count",
            "summary",
            "core_problem",
            "project_relevance",
            "note_path",
        ]
    )

    notes_root = build_paths()["notes_dir"]

    for paper in papers:
        stem = paper.get("source_file", "").replace(".tei.xml", "")
        matches = list(notes_root.rglob(f"{stem}.md"))
        note_path = matches[0] if matches else notes_root / "missing" / f"{stem}.md"
        sections = parse_note_sections(note_path)
        ws_papers.append(
            [
                paper.get("title", ""),
                paper.get("doi", ""),
                ", ".join(paper.get("authors", [])),
                len(paper.get("sections", [])),
                len(paper.get("references", [])),
                sections.get("一句话总结", ""),
                sections.get("核心问题", ""),
                sections.get("与本项目的关联", ""),
                str(note_path),
            ]
        )

    ws_queue = wb.create_sheet("queue")
    if queue_rows:
        ws_queue.append(list(queue_rows[0].keys()))
        for row in queue_rows:
            ws_queue.append([row[key] for key in row.keys()])
    else:
        ws_queue.append(["message"])
        ws_queue.append(["reading_queue.csv 尚未生成"])

    ws_stats = wb.create_sheet("stats")
    ws_stats.append(["metric", "value"])
    ws_stats.append(["generated_at", datetime.now().isoformat(timespec="seconds")])
    ws_stats.append(["seed_paper_count", len(papers)])
    ws_stats.append(["queue_count", len(queue_rows)])
    ws_stats.append(["concept_count", concept_map.get("stats", {}).get("concept_count", 0)])
    ws_stats.append(["concept_edge_count", concept_map.get("stats", {}).get("edge_count", 0)])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_summary(
    papers: list[dict],
    queue_rows: list[dict],
    concept_map: dict,
    download_rows: list[dict],
    output_path: Path,
) -> None:
    concept_counts = sorted(
        concept_map.get("concepts", []),
        key=lambda item: (-item.get("paper_count", 0), item.get("label", "")),
    )
    top_concepts = ", ".join(
        f"{item['label']}({item['paper_count']})" for item in concept_counts[:6]
    )

    must_read = [row for row in queue_rows if row.get("priority") == "must_read"]
    recommended = [row for row in queue_rows if row.get("priority") == "recommended"]
    downloaded_count = sum(
        1
        for row in download_rows
        if row.get("status") in {"downloaded", "already_downloaded"}
    )
    manual_count = sum(
        1 for row in download_rows if row.get("status") == "manual_download_needed"
    )
    round_one_count = max(0, len(papers) - 10)

    lines = [
        "# 文献综述框架（原型）",
        "",
        f"- 生成时间: {datetime.now().isoformat(timespec='seconds')}",
        f"- 已解析种子论文: {len(papers)} 篇",
        f"- 第一轮候选参考文献: {len(queue_rows)} 篇",
        "",
        "## 当前观察",
        "",
        "种子论文已经形成较清晰的三条主线：",
        "1. Agent/LLM 驱动的端到端 DFT 自动化框架，例如 DREAMS、TritonDFT、VASPilot、El Agente、Masgent。",
        "2. 作为底座的高通量与工作流基础设施，例如 atomate2，以及材料数据库、结构工具和 HPC 调度工具。",
        "3. 围绕可靠性的关键能力，包括参数收敛、错误恢复、并行资源配置、人类反馈与记忆机制。",
        "",
        "## 概念覆盖",
        "",
        f"当前概念图中出现频率最高的概念包括：{top_concepts or '待生成'}。",
        "",
        "## 第一轮阅读建议",
        "",
        f"- 必读候选: {len(must_read)} 篇",
        f"- 推荐候选: {len(recommended)} 篇",
        f"- round_1 已进入结构化结果: {round_one_count} 篇",
        f"- 当前下载成功/已存在: {downloaded_count} 篇",
        f"- 当前需人工下载: {manual_count} 篇",
        "",
    ]

    if queue_rows:
        lines.append("优先扩展的候选包括：")
        lines.append("")
        for row in queue_rows[:8]:
            lines.append(
                f"- {row['title']} | 分数 {row['score_total']} | 优先级 {row['priority']}"
            )
        lines.append("")

    lines.extend(
        [
            "## 当前不足",
            "",
            "- 当前已精修 4 篇核心论文；其余新增论文仍是初版笔记，实验设置、关键指标与局限性需要继续细化。",
            "- 第一轮队列仍是基于本地启发式评分，尚未接入 Semantic Scholar 反向引用和开放获取查询。",
            "- 自动下载阶段已经启动，但仍有多篇高优先级论文只能标记为 manual_download_needed。",
            "",
            "## 后续动作",
            "",
            "1. 继续精修 round_1 中与 Agent/DFT 最相关的新增论文笔记。",
            "2. 对当前 manual_download_needed 条目做人工补档或补充开放获取来源。",
            "3. 若继续扩展，则在现有 15 篇基础上进入下一轮下载、解析与评分。",
        ]
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines), encoding="utf-8")


def write_knowledge_gaps(concept_map: dict, output_path: Path) -> None:
    concept_labels = {item["label"] for item in concept_map.get("concepts", [])}

    gaps = [
        "# 知识空白与后续方向",
        "",
        "## 已经覆盖较多的方向",
        "",
        "- Agent/LLM 驱动 DFT 自动化",
        "- 工作流与 HPC 编排",
        "- 参数推断、收敛与错误恢复",
        "",
        "## 仍需重点补齐的方向",
        "",
    ]

    if "structure_modeling" in concept_labels:
        gaps.append("- 自动化建模已出现，但针对复杂表面、缺陷、溶剂、反应路径的程序化建模还不充分。")
    else:
        gaps.append("- 自动化建模方向覆盖不足，需要补结构生成、表面/缺陷建模、吸附位点搜索论文。")

    if "materials_db" in concept_labels:
        gaps.append("- 数据库与结构工具被频繁提及，但缺少与实际项目接口设计相关的工程论文。")
    else:
        gaps.append("- 材料数据库与结构工具接口覆盖不足。")

    gaps.extend(
        [
            "- 多 agent 通信协议、共享状态和长期记忆机制仍缺少针对 DFT 软件工程的系统比较。",
            "- 错误恢复目前多集中在收敛问题，针对作业调度失败、文件损坏、依赖缺失的恢复策略覆盖不足。",
            "- 面向普通用户的一键化交互与结果解释层几乎没有系统化总结。",
            "",
            "## 对本项目的直接启示",
            "",
            "- 后续系统设计要把结构建模、输入生成、执行调度、错误恢复、结果判读拆成稳定模块。",
            "- 需要提前设计可审计的执行日志和共享上下文，而不是仅靠聊天历史。",
        ]
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(gaps), encoding="utf-8")


def main() -> int:
    paths = build_paths()

    parser = argparse.ArgumentParser(description="生成汇总表、综述和知识空白文档")
    parser.add_argument(
        "--input-dirs",
        nargs="+",
        default=[str(paths["parsed_seed_info"])],
        help="结构化论文 JSON 目录列表",
    )
    parser.add_argument(
        "--queue-csv",
        default=str(paths["reading_queue_csv"]),
        help="待读队列 CSV 路径",
    )
    parser.add_argument(
        "--download-log",
        default=str(paths["download_log_csv"]),
        help="下载日志 CSV 路径",
    )
    parser.add_argument(
        "--concept-map",
        default=str(paths["concept_map_json"]),
        help="概念图 JSON 路径",
    )
    parser.add_argument(
        "--xlsx",
        default=str(paths["literature_review_xlsx"]),
        help="汇总表输出路径",
    )
    parser.add_argument(
        "--summary-md",
        default=str(paths["review_summary_md"]),
        help="综述输出路径",
    )
    parser.add_argument(
        "--gaps-md",
        default=str(paths["knowledge_gaps_md"]),
        help="知识空白输出路径",
    )
    args = parser.parse_args()

    papers = load_json_files([Path(item) for item in args.input_dirs])
    if not papers:
        print(f"未找到结构化论文 JSON: {args.input_dirs}")
        return 1

    queue_rows = load_csv_rows(Path(args.queue_csv))
    download_rows = load_csv_rows(Path(args.download_log))
    concept_path = Path(args.concept_map)
    concept_map = (
        json.loads(concept_path.read_text(encoding="utf-8"))
        if concept_path.exists()
        else {"stats": {}, "concepts": []}
    )

    build_workbook(papers, queue_rows, concept_map, Path(args.xlsx))
    write_summary(papers, queue_rows, concept_map, download_rows, Path(args.summary_md))
    write_knowledge_gaps(concept_map, Path(args.gaps_md))

    print(f"汇总表: {args.xlsx}")
    print(f"综述文档: {args.summary_md}")
    print(f"知识空白: {args.gaps_md}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
